"""
정산 결과를 '수식이 적용된' 엑셀(.xlsx)로 생성.

구성
  - 원본 시트: 원본_AXZ / 원본_SAP / 원본_카카오빌링 / 원본_현금영수증 / 원본_PG
  - 원본_AXZ 에는 보조 수식열(정산월여부/최종매출인식금액/구다음/대상매출)을 추가
  - 요약 시트: 요약_AXZ정산 / 요약_PG / 요약_현금영수증 / 요약_세금계산서 / 정산제외
    · AXZ정산·세금계산서·PG·현금영수증 원본합계는 SUMIFS/SUMIF 수식 → 원본 수정 시 자동 재계산
    · 현금영수증 시차보정·카카오 실제발행 등 순수 수식화가 어려운 값은 계산값으로 기입
"""
import io
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reconcile import _read_bytes, clean_id, run_reconciliation

# ---- 스타일 ----
HDR_FILL = PatternFill("solid", fgColor="1E293B")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="0F172A")
TOTAL_FILL = PatternFill("solid", fgColor="F1F5F9")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WON_FMT = "#,##0"


def _clean_val(v):
    if isinstance(v, float) and pd.isna(v):
        return None
    if v is pd.NaT:
        return None
    if isinstance(v, pd.Timestamp):
        return None if pd.isna(v) else v.to_pydatetime()
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return None if pd.isna(v) else float(v)
    return v


def _append_df(ws, df):
    """DataFrame 을 헤더+데이터로 기록. 반환: (헤더 리스트, 데이터행 수)."""
    headers = [str(c) for c in df.columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    n = 0
    for rec in df.itertuples(index=False, name=None):
        ws.append([_clean_val(v) for v in rec])
        n += 1
    return headers, n


def _colletter(headers, name):
    return get_column_letter(headers.index(name) + 1)


def _title(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT


def build_workbook(*, bill, sap, rec, axz, pg_files=None, target_month_str):
    pg_files = pg_files or []
    tm = target_month_str

    # ---------- 원본 로드 & 정제(계산 로직과 동일 규칙) ----------
    df_axz = _read_bytes(axz["bytes"], axz["name"]); df_axz.columns = df_axz.columns.str.strip()
    df_axz["거래일시"] = pd.to_datetime(df_axz["거래일시"], errors="coerce")
    df_axz["결제금액"] = pd.to_numeric(df_axz["결제금액"].astype(str).str.replace(",", ""), errors="coerce").fillna(0)
    df_axz["환불금액"] = pd.to_numeric(df_axz["환불금액"].astype(str).str.replace(",", ""), errors="coerce").fillna(0)
    for c in ["결제ID", "User ID", "사업자번호"]:
        if c in df_axz.columns:
            df_axz[c] = df_axz[c].apply(clean_id)
    for c in ["상품명", "비고", "결제수단", "세금유형", "결제상태"]:
        if c in df_axz.columns:
            df_axz[c] = df_axz[c].fillna("")

    df_sap = _read_bytes(sap["bytes"], sap["name"]); df_sap.columns = df_sap.columns.str.strip()
    df_sap = df_sap[~df_sap["거래처 1"].astype(str).str.contains("합계|합산|Total", na=False, case=False)].dropna(subset=["거래처 1"])
    df_sap["선수금"] = pd.to_numeric(df_sap["선수금"].astype(str).str.replace(",", ""), errors="coerce").fillna(0)
    df_sap["거래처 1"] = df_sap["거래처 1"].apply(clean_id)

    df_bill = _read_bytes(bill["bytes"], bill["name"]); df_bill.columns = df_bill.columns.str.strip()
    df_bill["결제금액"] = pd.to_numeric(df_bill["결제금액"].astype(str).str.replace(",", ""), errors="coerce").fillna(0)

    df_rec = _read_bytes(rec["bytes"], rec["name"]); df_rec.columns = df_rec.columns.str.strip()

    df_pg = (pd.concat([_read_bytes(pf["bytes"], pf["name"]) for pf in pg_files], ignore_index=True)
             if pg_files else pd.DataFrame())
    if not df_pg.empty:
        df_pg.columns = df_pg.columns.str.strip()

    # 계산값(수식화 어려운 부분 및 참고용) 확보
    res = run_reconciliation(bill=bill, sap=sap, rec=rec, axz=axz, pg_files=pg_files, target_month_str=tm)
    summ = res["summary"]

    wb = Workbook()
    wb.remove(wb.active)

    # ---------- 원본_AXZ (+ 보조 수식열) ----------
    ws = wb.create_sheet("원본_AXZ")
    headers, n = _append_df(ws, df_axz)
    last = n + 1  # 마지막 데이터 행
    L_date = _colletter(headers, "거래일시")
    L_pay = _colletter(headers, "결제금액")
    L_ref = _colletter(headers, "환불금액")
    L_method = _colletter(headers, "결제수단")
    L_tax = _colletter(headers, "세금유형")
    L_prod = _colletter(headers, "상품명")
    L_note = _colletter(headers, "비고")
    L_pid = _colletter(headers, "결제ID")
    L_biz = _colletter(headers, "사업자번호")

    base = len(headers)
    c_month = base + 1   # 정산월여부
    c_final = base + 2   # 최종매출인식금액
    c_dtext = base + 3   # 구다음(텍스트)
    c_dfin = base + 4    # 구다음(최종)
    c_tgt = base + 5     # 대상여부
    c_amt = base + 6     # 대상매출
    hdrs2 = ["정산월여부", "최종매출인식금액", "구다음(텍스트)", "구다음(최종)", "대상여부", "대상매출"]
    for i, h in enumerate(hdrs2):
        cc = ws.cell(row=1, column=base + 1 + i, value=h)
        cc.fill = HDR_FILL; cc.font = HDR_FONT; cc.alignment = Alignment(horizontal="center")

    Ldtext = get_column_letter(c_dtext)
    for r in range(2, last + 1):
        ws.cell(row=r, column=c_month,
                value=f'=IF(TEXT({L_date}{r},"YYYY-MM")="{tm}",1,0)')
        ws.cell(row=r, column=c_final,
                value=f'=IF({get_column_letter(c_month)}{r}=1,{L_pay}{r}-{L_ref}{r},-{L_ref}{r})')
        ws.cell(row=r, column=c_dtext,
                value=(f'=IF(OR(ISNUMBER(SEARCH("구다음메일",{L_prod}{r})),'
                       f'ISNUMBER(SEARCH("구다음결제",{L_note}{r}))),1,0)'))
        ws.cell(row=r, column=c_dfin,
                value=(f'=IF({get_column_letter(c_dtext)}{r}=1,1,'
                       f'IF(AND({L_pid}{r}<>"",'
                       f'SUMIFS(${Ldtext}$2:${Ldtext}${last},${L_pid}$2:${L_pid}${last},{L_pid}{r})>0),1,0))'))
        ws.cell(row=r, column=c_tgt,
                value=f'=IF({get_column_letter(c_dfin)}{r}=1,0,1)')
        ws.cell(row=r, column=c_amt,
                value=f'={get_column_letter(c_tgt)}{r}*{get_column_letter(c_final)}{r}')
        ws.cell(row=r, column=c_final).number_format = WON_FMT
        ws.cell(row=r, column=c_amt).number_format = WON_FMT

    L_final = get_column_letter(c_final)
    L_amt = get_column_letter(c_amt)
    R_amt = f"'원본_AXZ'!${L_amt}$2:${L_amt}${last}"
    R_final = f"'원본_AXZ'!${L_final}$2:${L_final}${last}"
    R_method = f"'원본_AXZ'!${L_method}$2:${L_method}${last}"
    R_tax = f"'원본_AXZ'!${L_tax}$2:${L_tax}${last}"
    R_biz = f"'원본_AXZ'!${L_biz}$2:${L_biz}${last}"

    # ---------- 원본_SAP / 카카오빌링 / 현금영수증 / PG ----------
    ws_sap = wb.create_sheet("원본_SAP")
    h_sap, n_sap = _append_df(ws_sap, df_sap)
    L_sbiz = _colletter(h_sap, "거래처 1"); L_sadv = _colletter(h_sap, "선수금")
    R_sbiz = f"'원본_SAP'!${L_sbiz}$2:${L_sbiz}${n_sap+1}"
    R_sadv = f"'원본_SAP'!${L_sadv}$2:${L_sadv}${n_sap+1}"

    ws_bill = wb.create_sheet("원본_카카오빌링")
    h_bill, n_bill = _append_df(ws_bill, df_bill)
    L_bcash = _colletter(h_bill, "캐시구분"); L_bpay = _colletter(h_bill, "결제금액")
    R_bcash = f"'원본_카카오빌링'!${L_bcash}$2:${L_bcash}${n_bill+1}"
    R_bpay = f"'원본_카카오빌링'!${L_bpay}$2:${L_bpay}${n_bill+1}"

    ws_rec = wb.create_sheet("원본_현금영수증")
    _append_df(ws_rec, df_rec)

    if not df_pg.empty:
        ws_pg = wb.create_sheet("원본_PG")
        _append_df(ws_pg, df_pg)

    # 고유값
    tax_types = [t for t in pd.unique(df_axz["세금유형"]) if str(t).strip()]
    methods = [m for m in pd.unique(df_axz["결제수단"]) if str(m).strip()]

    # ---------- 요약_AXZ정산 (SUMIFS 피벗, 대상매출 기준) ----------
    wsx = wb.create_sheet("요약_AXZ정산")
    _title(wsx, 1, f"AXZ 정산 요약 (정산월 {tm}, 구다음 제외 · 원본_AXZ 수식 참조)")
    hr = 3
    wsx.cell(row=hr, column=1, value="결제수단").font = BOLD
    for j, t in enumerate(tax_types):
        wsx.cell(row=hr, column=2 + j, value=t).font = BOLD
    wsx.cell(row=hr, column=2 + len(tax_types), value="합계").font = BOLD
    for i, m in enumerate(methods):
        r = hr + 1 + i
        wsx.cell(row=r, column=1, value=m)
        for j, t in enumerate(tax_types):
            col = 2 + j
            cl = get_column_letter(col)
            wsx.cell(row=r, column=col,
                     value=f'=SUMIFS({R_amt},{R_method},$A{r},{R_tax},{cl}${hr})').number_format = WON_FMT
        tot_c = 2 + len(tax_types)
        f = get_column_letter(2); l = get_column_letter(tot_c - 1)
        wsx.cell(row=r, column=tot_c, value=f"=SUM({f}{r}:{l}{r})").number_format = WON_FMT
    trow = hr + 1 + len(methods)
    wsx.cell(row=trow, column=1, value="합계").font = BOLD
    for j in range(len(tax_types) + 1):
        col = 2 + j; cl = get_column_letter(col)
        c = wsx.cell(row=trow, column=col, value=f"=SUM({cl}{hr+1}:{cl}{trow-1})")
        c.number_format = WON_FMT; c.font = BOLD
    for cell in wsx[trow]:
        cell.fill = TOTAL_FILL

    # ---------- 요약_세금계산서 (AXZ SUMIFS vs SAP SUMIF) ----------
    wst = wb.create_sheet("요약_세금계산서")
    _title(wst, 1, "세금계산서 대사 (AXZ 대상매출 vs SAP 선수금 · 수식)")
    biz_axz = df_axz[(df_axz["세금유형"] == "세금계산서")]["사업자번호"].apply(clean_id)
    biz_list = sorted(set([b for b in biz_axz if b]) | set([b for b in df_sap["거래처 1"] if b]))
    hr = 3
    for j, h in enumerate(["사업자번호", "AXZ(대상매출)", "SAP(선수금)", "차액"]):
        wst.cell(row=hr, column=1 + j, value=h).font = BOLD
    for i, b in enumerate(biz_list):
        r = hr + 1 + i
        wst.cell(row=r, column=1, value=b)
        wst.cell(row=r, column=2,
                 value=f'=SUMIFS({R_amt},{R_tax},"세금계산서",{R_biz},$A{r})').number_format = WON_FMT
        wst.cell(row=r, column=3, value=f'=SUMIF({R_sbiz},$A{r},{R_sadv})').number_format = WON_FMT
        wst.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = WON_FMT
    trow = hr + 1 + len(biz_list)
    wst.cell(row=trow, column=1, value="합계").font = BOLD
    for col in (2, 3, 4):
        cl = get_column_letter(col)
        c = wst.cell(row=trow, column=col, value=f"=SUM({cl}{hr+1}:{cl}{trow-1})")
        c.number_format = WON_FMT; c.font = BOLD
    for cell in wst[trow]:
        cell.fill = TOTAL_FILL

    # ---------- 요약_PG (AXZ 결제수단별 SUMIFS + 빌링 캐시구분별 SUMIF) ----------
    wsp = wb.create_sheet("요약_PG")
    _title(wsp, 1, "PG 요약 (AXZ 결제수단별 · 카카오빌링 캐시구분별 · 수식)")
    hr = 3
    wsp.cell(row=hr, column=1, value="[AXZ 결제수단별 대상매출]").font = BOLD
    for i, m in enumerate(methods):
        r = hr + 1 + i
        wsp.cell(row=r, column=1, value=m)
        wsp.cell(row=r, column=2, value=f'=SUMIFS({R_amt},{R_method},$A{r})').number_format = WON_FMT
    cashes = [c for c in pd.unique(df_bill["캐시구분"]) if str(c).strip()]
    hr2 = hr + 1 + len(methods) + 2
    wsp.cell(row=hr2, column=1, value="[카카오빌링 캐시구분별 결제금액]").font = BOLD
    for i, cash in enumerate(cashes):
        r = hr2 + 1 + i
        wsp.cell(row=r, column=1, value=cash)
        wsp.cell(row=r, column=2, value=f'=SUMIF({R_bcash},$A{r},{R_bpay})').number_format = WON_FMT

    # 대사 결과(참고, 계산값) — 매핑 병합 결과
    hr3 = hr2 + 1 + len(cashes) + 2
    wsp.cell(row=hr3, column=1, value="[대사 결과(참고) — 매핑 병합]").font = BOLD
    for j, h in enumerate(["결제유형", "최종매출인식금액", "결제금액", "차액"]):
        wsp.cell(row=hr3 + 1, column=1 + j, value=h).font = BOLD
    for i, row in enumerate(res["pg_summary"]):
        r = hr3 + 2 + i
        wsp.cell(row=r, column=1, value=row.get("결제유형"))
        for j, k in enumerate(["최종매출인식금액", "결제금액", "차액"]):
            wsp.cell(row=r, column=2 + j, value=row.get(k)).number_format = WON_FMT

    # ---------- 요약_현금영수증 ----------
    wsr = wb.create_sheet("요약_현금영수증")
    _title(wsr, 1, "현금영수증 대사")

    # 소거/합산 산출근거: 아래 상세표(사유=F, AXZ=C, 카카오원본=D)를 SUMIF 로 참조
    err = pd.DataFrame(res["err_rec"])
    DET_HDR = 13          # 상세표 헤더 행
    DET_START = DET_HDR + 1
    n_err = len(err)
    DET_END = DET_START + n_err - 1 if n_err else DET_START
    rng_reason = f"F{DET_START}:F{DET_END}"
    rng_axz = f"C{DET_START}:C{DET_END}"
    rng_kakao = f"D{DET_START}:D{DET_END}"
    if n_err:
        sogo_val = f'=-SUMIF({rng_reason},"익월 발행 예정건 (소거)",{rng_axz})'
        hapsan_val = f'=SUMIF({rng_reason},"전월 말일 결제건 (합산)",{rng_kakao})'
    else:  # 상세 없음 → 계산값 폴백
        sogo_val = -summ["end_of_month_sum"]
        hapsan_val = summ["prev_month_end_sum"]

    # AXZ 원본합계는 수식(현금영수증+자진발급, 최종매출인식금액 기준 · 구다음 포함 = 원본 로직과 동일)
    rows = [
        ("AXZ 현금영수증 원본합계(수식)",
         f'=SUMIFS({R_final},{R_tax},"현금영수증")+SUMIFS({R_final},{R_tax},"자진발급")', True),
        ("(-) 익월 발행 예정건 (소거)", sogo_val, False),
        ("(+) 전월 말일 결제건 (합산)", hapsan_val, False),
        ("보정된 AXZ 대상 총액", None, "calc_adj"),
        ("카카오 실제 발행액", summ["total_k_rec"], False),
        ("최종 발행 차액", None, "calc_diff"),
    ]
    rr = 3
    for label, val, kind in rows:
        wsr.cell(row=rr, column=1, value=label)
        cell = wsr.cell(row=rr, column=2)
        if kind is True:
            cell.value = val
        elif kind == "calc_adj":
            cell.value = f"=B{rr-3}+B{rr-2}+B{rr-1}"  # 원본합계 + (-익월) + (+전월)
        elif kind == "calc_diff":
            cell.value = f"=B{rr-2}-B{rr-1}"  # 보정된대상 - 카카오발행
        else:
            cell.value = val
        cell.number_format = WON_FMT
        rr += 1
    wsr.cell(row=rr + 1, column=1,
             value="※ 소거/합산은 아래 상세의 SUMIF 근거식입니다. 카카오 실제발행(중복제거·PG실패 제외)은 계산값입니다.").font = Font(size=9, color="64748B", italic=True)

    # ---- 소거/합산 산출근거 상세표 ----
    _title(wsr, DET_HDR - 1, "■ 소거 / 합산 산출근거 (현금영수증 대사 상세)")
    det_cols = ["거래일시", "계정 ID", "AXZ", "카카오원본", "차액", "사유"]
    for j, h in enumerate(det_cols):
        wsr.cell(row=DET_HDR, column=1 + j, value=h).font = BOLD
    if n_err:
        for i, (_, r) in enumerate(err.iterrows()):
            ri = DET_START + i
            wsr.cell(row=ri, column=1, value=r.get("거래일시"))
            wsr.cell(row=ri, column=2, value=r.get("계정 ID"))
            for cj, key in ((3, "AXZ"), (4, "카카오원본"), (5, "차액")):
                c = wsr.cell(row=ri, column=cj, value=r.get(key))
                c.number_format = WON_FMT
            wsr.cell(row=ri, column=6, value=r.get("사유"))
    else:
        wsr.cell(row=DET_START, column=1, value="상세 불일치 없음")

    # ---------- 정산제외 (구다음) ----------
    wse = wb.create_sheet("정산제외")
    _title(wse, 1, "정산 제외 (구다음메일 결제건)")
    old = pd.DataFrame(res["old_daum"])
    _title(wse, 2, "AXZ 결제내역 (비고/상품명에 '구다음' 포함)")
    if not old.empty:
        _append_df_at(wse, old, start_row=3)
        next_row = 3 + len(old) + 3
    else:
        wse.cell(row=3, column=1, value="제외 항목 없음")
        next_row = 5

    old_rec = pd.DataFrame(res.get("old_daum_rec", []))
    _title(wse, next_row, "현금영수증 (채널 = '다음메일')")
    if not old_rec.empty:
        _append_df_at(wse, old_rec, start_row=next_row + 1)
    else:
        wse.cell(row=next_row + 1, column=1, value="제외 항목 없음")

    _autofit(wb)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _append_df_at(ws, df, start_row):
    headers = [str(c) for c in df.columns]
    for j, h in enumerate(headers):
        cc = ws.cell(row=start_row, column=1 + j, value=h)
        cc.fill = HDR_FILL; cc.font = HDR_FONT; cc.alignment = Alignment(horizontal="center")
    for i, rec in enumerate(df.itertuples(index=False, name=None)):
        for j, v in enumerate(rec):
            ws.cell(row=start_row + 1 + i, column=1 + j, value=_clean_val(v))


def _autofit(wb):
    for ws in wb.worksheets:
        widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col = cell.column_letter
                ln = len(str(cell.value))
                widths[col] = min(max(widths.get(col, 10), ln + 2), 42)
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
