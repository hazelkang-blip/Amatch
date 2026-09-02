"""세금계산서 발행 ERP 원본(더존/SAP) 대사 검증.

더존 양식: '품의내역' 괄호 안 숫자 = 사업자발급번호, '장부금액'(= '거래금액') = 발행금액.
          판매금액(공급가) 열은 없고, 취소분은 음수로 노출된다. 마지막 총계행은 제외.
실행: python tests/test_erp_douzone.py
"""
import io, os, sys
import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "api"))
from reconcile import normalize_erp, run_reconciliation  # noqa
from excel_export import build_workbook  # noqa

def xls(df):
    b = io.BytesIO(); df.to_excel(b, index=False); return b.getvalue()

def dz_row(no, biz, amt, day="2025-07-10"):
    return {"No": no, "연동시스템": "프리미엄메일", "메뉴구분": "프리미엄메일",
            "업무구분": "콘텐츠매출_기타", "회계일자": day, "모듈관리번호": f"axzpm-{no}",
            "수정차수": "1", "품의내역": f"프리미엄메일_202507 ({biz})", "환종": "KRW",
            "장부금액": amt, "거래금액": amt, "처리구분": "이관완료", "전표번호": None}

def axz_tax(biz, pay, refund=0, dt="2025-07-10 10:00", req=None):
    row = {"거래일시": dt, "결제수단": "신용카드", "결제상태": "결제완료", "결제금액": pay,
           "환불금액": refund, "User ID": biz, "결제ID": f"P{biz}", "사업자번호": biz,
           "세금유형": "세금계산서", "상품명": "메일", "비고": ""}
    if req is not None:                    # 추후 추가될 '세금계산서 신청일자' 열
        row["세금계산서 신청일자"] = req
    return row

bill = pd.DataFrame([{"캐시구분": "신용카드", "결제금액": 0}])
rec = pd.DataFrame([{"계정ID": "z", "요청금액": 0, "전송유형": "결제", "승인일시": "2025-07-10 10:00"}])

# 총계행(품의내역 공란 + 금액만 있는 마지막 행) 포함
douzone = pd.DataFrame([
    dz_row(1, "1112233444", 34900),
    dz_row(2, "2223344555", 34900),
    dz_row(3, "3334455666", -32797),           # 거래취소(부분환불) → 음수
    {"No": " ", "연동시스템": "", "메뉴구분": "", "업무구분": "", "회계일자": "",
     "모듈관리번호": "", "수정차수": "", "품의내역": "", "환종": "",
     "장부금액": 37003, "거래금액": 37003, "처리구분": "", "전표번호": ""},
])

# 1) normalize_erp 단위 검증
n, fmt = normalize_erp(douzone)
assert fmt == "더존", fmt
assert len(n) == 3, f"총계행이 제거되지 않음: {len(n)}"
assert list(n["사업자번호"]) == ["1112233444", "2223344555", "3334455666"], list(n["사업자번호"])
assert n["발행금액"].sum() == 37003, n["발행금액"].sum()
assert list(n["신청일자"].dt.strftime("%Y-%m-%d")) == ["2025-07-10"] * 3, list(n["신청일자"])

# SAP 양식도 그대로 인식되는지
n2, fmt2 = normalize_erp(pd.DataFrame([{"거래처 1": "1112233444", "선수금": 34900},
                                       {"거래처 1": None, "선수금": 34900}]))
assert fmt2 == "SAP" and len(n2) == 1 and n2["발행금액"].sum() == 34900

# 2) 대사 결과: 1112233444 일치 / 2223344555 는 AXZ 없음 / 3334455666 부분환불 일치
axz = pd.DataFrame([
    axz_tax("1112233444", 34900),
    axz_tax("3334455666", 34900, refund=67697),   # 34900-67697 = -32797
])
res = run_reconciliation(
    bill={"bytes": xls(bill), "name": "b.xlsx"},
    sap={"bytes": xls(douzone), "name": "GLDDOC00200.xlsx"},
    rec={"bytes": xls(rec), "name": "r.xlsx"},
    axz={"bytes": xls(axz), "name": "a.xlsx"},
    target_month_str="2025-07")

assert res["erp_format"] == "더존", res["erp_format"]
err = {r["사업자번호"]: r for r in res["err_tax"]}
assert set(err) == {"2223344555"}, f"불일치 목록이 예상과 다름: {set(err)}"
assert err["2223344555"]["더존 원본"] == 34900
assert err["2223344555"]["차액"] == -34900
# 신청일자: 더존 회계일자만 있고 DAUM 결제내역에 신청일자 열이 없으면 ERP 열만 노출
assert res["err_tax_columns"] == ["사업자번호", "신청일자(더존)", "최종매출인식금액", "더존 원본", "차액"], res["err_tax_columns"]
assert err["2223344555"]["신청일자(더존)"] == "2025-07-10"

# 2-b) DAUM 결제내역에 '세금계산서 신청일자' 열이 생기면 나란히 비교
axz_req = pd.DataFrame([
    axz_tax("1112233444", 34900, req="2025-07-09"),
    axz_tax("3334455666", 34900, refund=67697, req="2025-07-11"),
])
res2 = run_reconciliation(
    bill={"bytes": xls(bill), "name": "b.xlsx"},
    sap={"bytes": xls(douzone), "name": "GLDDOC00200.xlsx"},
    rec={"bytes": xls(rec), "name": "r.xlsx"},
    axz={"bytes": xls(axz_req), "name": "a.xlsx"},
    target_month_str="2025-07")
assert res2["err_tax_columns"] == ["사업자번호", "신청일자(DAUM)", "신청일자(더존)",
                                   "최종매출인식금액", "더존 원본", "차액"], res2["err_tax_columns"]
e2 = {r["사업자번호"]: r for r in res2["err_tax"]}
# 더존에만 있는 건은 DAUM 신청일자가 '-'
assert e2["2223344555"]["신청일자(DAUM)"] == "-" and e2["2223344555"]["신청일자(더존)"] == "2025-07-10"
assert res2["err_tax_money_columns"] == ["최종매출인식금액", "더존 원본", "차액"]

# 3) 엑셀 export 도 더존 양식으로 생성되는지
xlsx = build_workbook(
    bill={"bytes": xls(bill), "name": "b.xlsx"},
    sap={"bytes": xls(douzone), "name": "GLDDOC00200.xlsx"},
    rec={"bytes": xls(rec), "name": "r.xlsx"},
    axz={"bytes": xls(axz_req), "name": "a.xlsx"},
    target_month_str="2025-07")
assert len(xlsx) > 1000
import openpyxl
wb = openpyxl.load_workbook(io.BytesIO(xlsx))
assert "원본_더존" in wb.sheetnames, wb.sheetnames
assert "원본_DAUM" in wb.sheetnames and "요약_DAUM정산" in wb.sheetnames, wb.sheetnames
hdr = [c.value for c in wb["요약_세금계산서"][3]]
assert hdr[:6] == ["사업자번호", "신청일자(DAUM)", "신청일자(더존)",
                   "DAUM(대상매출)", "더존(발행금액)", "차액"], hdr

print("PASS: 더존/SAP ERP 양식 대사 + 신청일자 참고열 + DAUM 표기 정상")
